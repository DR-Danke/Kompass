"""Supplier service for managing Chinese supplier business logic.

This service provides the business logic layer between API routes and the
repository layer for supplier CRUD operations, filtering, search, and
business rule enforcement.
"""

import io
import math
import re
from typing import Any, Dict, List, Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.kompass_dto import (
    PaginationDTO,
    SupplierCertificationSummaryDTO,
    SupplierCreateDTO,
    SupplierFromCardResultDTO,
    SupplierListResponseDTO,
    SupplierPipelineResponseDTO,
    SupplierPipelineStatus,
    SupplierPipelineSummaryDTO,
    SupplierResponseDTO,
    SupplierStatus,
    SupplierUpdateDTO,
    SupplierWithProductCountDTO,
)
from app.repository.kompass_repository import supplier_repository
from app.services.business_card_service import business_card_service


class SupplierService:
    """Handles supplier business logic including validation and CRUD operations."""

    # Email validation regex pattern
    EMAIL_PATTERN = re.compile(
        r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
    )

    def _validate_email(self, email: str) -> bool:
        """Validate email format.

        Args:
            email: Email address to validate

        Returns:
            True if valid, False otherwise
        """
        if not email:
            return True  # Empty email is allowed (optional field)
        return bool(self.EMAIL_PATTERN.match(email))

    def _normalize_wechat_id(self, wechat_id: Optional[str]) -> Optional[str]:
        """Normalize WeChat ID by converting to lowercase and stripping whitespace.

        Note: contact_phone field can be used to store WeChat IDs.

        Args:
            wechat_id: WeChat ID to normalize

        Returns:
            Normalized WeChat ID or None
        """
        if not wechat_id:
            return wechat_id
        return wechat_id.strip().lower()

    def create_supplier(self, request: SupplierCreateDTO) -> SupplierResponseDTO:
        """Create a new supplier with validation.

        Args:
            request: Supplier creation data

        Returns:
            Created supplier response

        Raises:
            ValueError: If validation fails or creation fails
        """
        # Validate email format (Pydantic EmailStr already validates,
        # but we add extra validation for safety)
        if request.contact_email and not self._validate_email(
            str(request.contact_email)
        ):
            raise ValueError("Invalid email format")

        # Normalize contact_phone (can be used for WeChat ID)
        contact_phone = self._normalize_wechat_id(request.contact_phone)

        # Create supplier via repository
        result = supplier_repository.create(
            name=request.name,
            code=request.code,
            status=request.status.value,
            contact_name=request.contact_name,
            contact_email=str(request.contact_email) if request.contact_email else None,
            contact_phone=contact_phone,
            address=request.address,
            city=request.city,
            country=request.country,
            website=request.website,
            notes=request.notes,
            wechat_id=request.wechat_id,
        )

        if not result:
            print("ERROR [SupplierService]: Failed to create supplier")
            raise ValueError("Failed to create supplier")

        print(f"INFO [SupplierService]: Created supplier {result['id']}")
        return SupplierResponseDTO(**result)

    def get_supplier(self, supplier_id: UUID) -> Optional[SupplierResponseDTO]:
        """Get a supplier by ID with product count.

        Args:
            supplier_id: UUID of the supplier

        Returns:
            Supplier response with product count, or None if not found
        """
        result = supplier_repository.get_by_id(supplier_id)

        if not result:
            print(f"INFO [SupplierService]: Supplier {supplier_id} not found")
            return None

        print(f"INFO [SupplierService]: Retrieved supplier {supplier_id}")
        return SupplierResponseDTO(**result)

    def get_supplier_with_product_count(
        self, supplier_id: UUID
    ) -> Optional[dict]:
        """Get a supplier by ID with product count.

        Args:
            supplier_id: UUID of the supplier

        Returns:
            Supplier dict with product_count field, or None if not found
        """
        result = supplier_repository.get_by_id(supplier_id)

        if not result:
            print(f"INFO [SupplierService]: Supplier {supplier_id} not found")
            return None

        # Get product count for this supplier
        product_count = supplier_repository.count_products_by_supplier(supplier_id)
        result["product_count"] = product_count

        print(f"INFO [SupplierService]: Retrieved supplier {supplier_id}")
        return result

    def list_suppliers(
        self,
        status: Optional[SupplierStatus] = None,
        country: Optional[str] = None,
        has_products: Optional[bool] = None,
        certification_status: Optional[str] = None,
        pipeline_status: Optional[str] = None,
        source: Optional[str] = None,
        search: Optional[str] = None,
        sort_by: str = "name",
        sort_order: str = "asc",
        page: int = 1,
        limit: int = 20,
    ) -> SupplierListResponseDTO:
        """List suppliers with filtering and pagination.

        Args:
            status: Filter by supplier status
            country: Filter by country
            has_products: Filter by whether supplier has products
            certification_status: Filter by certification status
            pipeline_status: Filter by pipeline status
            search: Search query for name, email, phone, code
            sort_by: Field to sort by (default: name)
            sort_order: Sort direction (asc, desc)
            page: Page number (1-indexed)
            limit: Items per page

        Returns:
            Paginated list of suppliers
        """
        # Use extended get_all with filters
        items, total = supplier_repository.get_all_with_filters(
            page=page,
            limit=limit,
            status=status.value if status else None,
            country=country,
            has_products=has_products,
            certification_status=certification_status,
            pipeline_status=pipeline_status,
            source=source,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
        )

        pages = math.ceil(total / limit) if total > 0 else 0

        supplier_responses = [SupplierResponseDTO(**item) for item in items]

        print(
            f"INFO [SupplierService]: Listed {len(supplier_responses)} suppliers "
            f"(page {page}/{pages})"
        )

        return SupplierListResponseDTO(
            items=supplier_responses,
            pagination=PaginationDTO(
                page=page,
                limit=limit,
                total=total,
                pages=pages,
            ),
        )

    def update_supplier(
        self, supplier_id: UUID, request: SupplierUpdateDTO
    ) -> Optional[SupplierResponseDTO]:
        """Update a supplier with validation.

        Args:
            supplier_id: UUID of the supplier to update
            request: Update data

        Returns:
            Updated supplier response, or None if not found

        Raises:
            ValueError: If validation fails
        """
        # Check if supplier exists
        existing = supplier_repository.get_by_id(supplier_id)
        if not existing:
            return None

        # Validate email format if provided
        if request.contact_email and not self._validate_email(
            str(request.contact_email)
        ):
            raise ValueError("Invalid email format")

        # Normalize contact_phone if provided
        contact_phone = self._normalize_wechat_id(request.contact_phone)

        # Build update kwargs, only including non-None values
        update_kwargs = {}
        if request.name is not None:
            update_kwargs["name"] = request.name
        if request.code is not None:
            update_kwargs["code"] = request.code
        if request.status is not None:
            update_kwargs["status"] = request.status.value
        if request.contact_name is not None:
            update_kwargs["contact_name"] = request.contact_name
        if request.contact_email is not None:
            update_kwargs["contact_email"] = str(request.contact_email)
        if contact_phone is not None:
            update_kwargs["contact_phone"] = contact_phone
        if request.address is not None:
            update_kwargs["address"] = request.address
        if request.city is not None:
            update_kwargs["city"] = request.city
        if request.country is not None:
            update_kwargs["country"] = request.country
        if request.website is not None:
            update_kwargs["website"] = request.website
        if request.notes is not None:
            update_kwargs["notes"] = request.notes
        if request.wechat_id is not None:
            update_kwargs["wechat_id"] = request.wechat_id

        result = supplier_repository.update(supplier_id, **update_kwargs)

        if not result:
            print(f"ERROR [SupplierService]: Failed to update supplier {supplier_id}")
            raise ValueError("Failed to update supplier")

        print(f"INFO [SupplierService]: Updated supplier {supplier_id}")
        return SupplierResponseDTO(**result)

    def delete_supplier(self, supplier_id: UUID) -> Optional[Dict[str, Any]]:
        """Hard delete a supplier and all associated data.

        Permanently removes the supplier and cascades deletion to products,
        audits, product images, product tags, and portfolio items.

        Args:
            supplier_id: UUID of the supplier to delete

        Returns:
            Dict with deletion counts, or None if supplier not found
        """
        # Check if supplier exists
        existing = supplier_repository.get_by_id(supplier_id)
        if not existing:
            return None

        result = supplier_repository.delete(supplier_id)

        if result:
            print(
                f"INFO [SupplierService]: Hard deleted supplier {supplier_id} "
                f"(products: {result['products_deleted']}, audits: {result['audits_deleted']})"
            )

        return result

    def get_delete_preview(self, supplier_id: UUID) -> Optional[Dict[str, Any]]:
        """Get preview of what would be deleted for a supplier.

        Args:
            supplier_id: UUID of the supplier

        Returns:
            Dict with supplier_name, products_count, audits_count, or None if not found
        """
        return supplier_repository.get_delete_preview(supplier_id)

    def search_suppliers(self, query: str) -> List[SupplierResponseDTO]:
        """Search suppliers by name, email, or contact phone (WeChat).

        Args:
            query: Search query string

        Returns:
            List of matching suppliers (max 50)
        """
        # Skip if query is empty or too short
        if not query or len(query.strip()) < 2:
            return []

        query = query.strip()

        # Use repository search with extended fields
        items = supplier_repository.search(query=query, limit=50)

        print(f"INFO [SupplierService]: Search for '{query}' returned {len(items)} results")

        return [SupplierResponseDTO(**item) for item in items]

    def list_certified_suppliers(
        self,
        grade: Optional[str] = None,
        page: int = 1,
        limit: int = 20,
    ) -> SupplierListResponseDTO:
        """List only certified suppliers (certified_a, certified_b, certified_c).

        Args:
            grade: Optional filter by grade (A, B, C)
            page: Page number (1-indexed)
            limit: Items per page

        Returns:
            Paginated list of certified suppliers

        Raises:
            ValueError: If invalid grade is provided
        """
        # Validate grade if provided
        if grade and grade.upper() not in ["A", "B", "C"]:
            raise ValueError(
                f"Invalid grade: {grade}. Must be one of: A, B, C"
            )

        items, total = supplier_repository.get_by_certification_status(
            grade=grade,
            page=page,
            limit=limit,
        )

        pages = math.ceil(total / limit) if total > 0 else 0

        supplier_responses = [SupplierResponseDTO(**item) for item in items]

        print(
            f"INFO [SupplierService]: Listed {len(supplier_responses)} certified suppliers "
            f"(grade={grade}, page {page}/{pages})"
        )

        return SupplierListResponseDTO(
            items=supplier_responses,
            pagination=PaginationDTO(
                page=page,
                limit=limit,
                total=total,
                pages=pages,
            ),
        )

    def list_suppliers_by_pipeline(
        self,
        pipeline_status: SupplierPipelineStatus,
        page: int = 1,
        limit: int = 20,
    ) -> SupplierListResponseDTO:
        """List suppliers filtered by pipeline status.

        Args:
            pipeline_status: Pipeline status to filter by
            page: Page number (1-indexed)
            limit: Items per page

        Returns:
            Paginated list of suppliers
        """
        items, total = supplier_repository.get_by_pipeline_status(
            pipeline_status=pipeline_status.value,
            page=page,
            limit=limit,
        )

        pages = math.ceil(total / limit) if total > 0 else 0

        supplier_responses = [SupplierResponseDTO(**item) for item in items]

        print(
            f"INFO [SupplierService]: Listed {len(supplier_responses)} suppliers "
            f"with pipeline_status={pipeline_status.value} (page {page}/{pages})"
        )

        return SupplierListResponseDTO(
            items=supplier_responses,
            pagination=PaginationDTO(
                page=page,
                limit=limit,
                total=total,
                pages=pages,
            ),
        )

    def update_pipeline_status(
        self,
        supplier_id: UUID,
        pipeline_status: SupplierPipelineStatus,
    ) -> Optional[SupplierResponseDTO]:
        """Update the pipeline status of a supplier.

        Args:
            supplier_id: UUID of the supplier
            pipeline_status: New pipeline status

        Returns:
            Updated supplier response, or None if not found
        """
        # Check if supplier exists
        existing = supplier_repository.get_by_id(supplier_id)
        if not existing:
            return None

        result = supplier_repository.update_pipeline_status(
            supplier_id=supplier_id,
            pipeline_status=pipeline_status.value,
        )

        if not result:
            print(f"ERROR [SupplierService]: Failed to update pipeline status for {supplier_id}")
            return None

        print(
            f"INFO [SupplierService]: Updated pipeline status for {supplier_id} "
            f"to {pipeline_status.value}"
        )
        return SupplierResponseDTO(**result)

    def get_certification_summary(
        self,
        supplier_id: UUID,
    ) -> Optional[SupplierCertificationSummaryDTO]:
        """Get certification summary for a supplier including latest audit info.

        Args:
            supplier_id: UUID of the supplier

        Returns:
            Certification summary with audit details, or None if not found
        """
        result = supplier_repository.get_with_certification_details(supplier_id)

        if not result:
            print(f"INFO [SupplierService]: Supplier {supplier_id} not found")
            return None

        print(f"INFO [SupplierService]: Retrieved certification summary for {supplier_id}")
        return SupplierCertificationSummaryDTO(**result)

    def get_pipeline_summary(self) -> SupplierPipelineSummaryDTO:
        """Get counts of suppliers grouped by pipeline status.

        Returns:
            Pipeline summary with counts for each status
        """
        result = supplier_repository.get_pipeline_summary()
        print("INFO [SupplierService]: Retrieved pipeline summary")
        return SupplierPipelineSummaryDTO(**result)

    def get_pipeline(self) -> SupplierPipelineResponseDTO:
        """Get all suppliers grouped by pipeline status for Kanban view.

        Returns:
            Suppliers grouped by pipeline status
        """
        result = supplier_repository.get_all_grouped_by_pipeline()

        # Convert dict data to DTOs
        pipeline_response = SupplierPipelineResponseDTO(
            contacted=[SupplierWithProductCountDTO(**s) for s in result.get("contacted", [])],
            potential=[SupplierWithProductCountDTO(**s) for s in result.get("potential", [])],
            quoted=[SupplierWithProductCountDTO(**s) for s in result.get("quoted", [])],
            certified=[SupplierWithProductCountDTO(**s) for s in result.get("certified", [])],
            active=[SupplierWithProductCountDTO(**s) for s in result.get("active", [])],
            inactive=[SupplierWithProductCountDTO(**s) for s in result.get("inactive", [])],
        )

        total_count = sum(
            len(result.get(status, []))
            for status in ["contacted", "potential", "quoted", "certified", "active", "inactive"]
        )
        print(f"INFO [SupplierService]: Retrieved pipeline with {total_count} total suppliers")
        return pipeline_response

    def create_supplier_from_card(self, capture_id: UUID) -> SupplierFromCardResultDTO:
        """Create a supplier from an extracted business card capture.

        Args:
            capture_id: UUID of the business card capture

        Returns:
            SupplierFromCardResultDTO with creation result

        Raises:
            ValueError: If capture not found, invalid status, or missing data
        """
        # 1. Retrieve capture
        capture = business_card_service.get_capture(capture_id)

        # 2. Validate status
        if capture["status"] != "extracted":
            raise ValueError(
                f"Cannot create supplier from capture with status '{capture['status']}'. "
                "Only 'extracted' captures can be used."
            )

        # 3. Check if already linked to a supplier
        if capture.get("supplier_id"):
            raise ValueError("This capture is already linked to a supplier")

        # 4. Extract fields
        company_name = capture.get("company_name")
        contact_name = capture.get("contact_name")
        contact_email = capture.get("contact_email")
        contact_phone = capture.get("contact_phone")
        contact_wechat = capture.get("contact_wechat")
        address = capture.get("address")
        website = capture.get("website")
        fair_name = capture.get("fair_name")

        # 5. Determine supplier name
        supplier_name = company_name or contact_name
        if not supplier_name:
            raise ValueError("No company or contact name extracted from business card")

        # 6. Duplicate detection
        dup = supplier_repository.find_duplicate_supplier(
            email=contact_email, phone=contact_phone
        )

        if dup:
            print(
                f"INFO [SupplierService]: Duplicate supplier detected for capture {capture_id}: "
                f"existing supplier {dup['id']}"
            )
            # Update capture status to rejected (duplicate)
            business_card_service.update_capture(capture_id, {"status": "rejected"})
            return SupplierFromCardResultDTO(
                success=False,
                is_duplicate=True,
                duplicate_supplier_id=dup["id"],
                duplicate_supplier_name=dup["name"],
                message="Proveedor duplicado detectado",
            )

        # 7. Validate email if present
        if contact_email and not self._validate_email(contact_email):
            contact_email = None  # Skip invalid email rather than failing

        # 8. Create supplier with trade fair metadata
        result = supplier_repository.create_with_trade_fair_metadata(
            name=supplier_name,
            contact_name=contact_name,
            contact_email=contact_email,
            contact_phone=contact_phone,
            address=address,
            country="China",
            website=website,
            pipeline_status="contacted",
            source="trade_fair",
            fair_name=fair_name,
            capture_date=capture.get("created_at"),
            wechat_id=contact_wechat,
        )

        if not result:
            raise ValueError("Failed to create supplier from business card")

        # 9. Link capture to supplier and update status
        business_card_service.update_capture(
            capture_id,
            {"supplier_id": str(result["id"]), "status": "confirmed"},
        )

        print(
            f"INFO [SupplierService]: Created supplier {result['id']} from capture {capture_id}"
        )

        # Send introduction email if setting enabled and supplier has an email
        from app.config.settings import get_settings

        settings = get_settings()
        email_sent = False
        email_error = None
        no_email_address = not bool(contact_email)

        if settings.AUTO_SEND_CARD_EMAIL:
            if contact_email:
                try:
                    from app.services.email_service import email_service

                    email_result = email_service.send_supplier_introduction(
                        supplier_name=supplier_name,
                        supplier_email=contact_email,
                        fair_name=fair_name or "the trade fair",
                    )
                    email_sent = email_result.success
                    print(
                        f"INFO [SupplierService]: Introduction email sent to {contact_email}: "
                        f"success={email_result.success}, mock={email_result.mock_mode}"
                    )
                except Exception as e:
                    email_error = str(e)
                    print(f"WARN [SupplierService]: Failed to send introduction email: {e}")
            else:
                print(f"INFO [SupplierService]: No email address for supplier {result['id']}, skipping auto-email")
        else:
            print(f"INFO [SupplierService]: AUTO_SEND_CARD_EMAIL disabled, skipping auto-email for {result['id']}")

        return SupplierFromCardResultDTO(
            success=True,
            supplier_id=result["id"],
            supplier_name=result["name"],
            message="Proveedor creado exitosamente",
            email_sent=email_sent,
            email_error=email_error,
            no_email_address=no_email_address,
        )

    def get_outreach_templates(self) -> list:
        """Return available outreach message templates.

        Returns:
            List of template metadata dicts
        """
        from app.services.email_service import email_service

        return email_service.get_templates()

    def send_outreach(
        self,
        supplier_id: UUID,
        template: str,
        channels: List[str],
        custom_message: Optional[str] = None,
    ) -> dict:
        """Send outreach messages to a supplier via selected channels.

        Args:
            supplier_id: UUID of the supplier
            template: Template key from OUTREACH_TEMPLATES
            channels: List of channels ('email', 'wechat')
            custom_message: Optional custom message override

        Returns:
            Dict with email_sent, wechat_sent, message, outreach_status

        Raises:
            ValueError: If supplier not found, invalid template, or invalid channels
        """
        from app.services.email_service import OUTREACH_TEMPLATES, email_service
        from app.services.wechat_service import wechat_service

        # Validate supplier exists
        supplier = supplier_repository.get_by_id_extended(supplier_id)
        if not supplier:
            raise ValueError("Supplier not found")

        # Validate template
        if template not in OUTREACH_TEMPLATES:
            valid_keys = ", ".join(OUTREACH_TEMPLATES.keys())
            raise ValueError(f"Invalid template: {template}. Must be one of: {valid_keys}")

        # Validate channels
        valid_channels = {"email", "wechat"}
        if not channels or not set(channels).issubset(valid_channels):
            raise ValueError("Channels must contain 'email' and/or 'wechat'")

        email_sent = False
        wechat_sent = False
        mock_mode = False
        messages = []

        # Send via email
        if "email" in channels:
            contact_email = supplier.get("contact_email")
            if contact_email:
                result = email_service.send_template_email(
                    supplier=supplier,
                    template_name=template,
                    custom_message=custom_message,
                )
                email_sent = result.success
                mock_mode = result.mock_mode
                messages.append(f"Email: {result.message}")
            else:
                messages.append("Email: No contact email available")

        # Send via WeChat
        if "wechat" in channels:
            wechat_id = supplier.get("wechat_id")
            if wechat_id:
                result = wechat_service.send_template_message(
                    supplier=supplier,
                    template_name=template,
                    custom_message=custom_message,
                )
                wechat_sent = result.success
                messages.append(f"WeChat: {result.message}")
            else:
                messages.append("WeChat: No WeChat ID available")

        # Update outreach status if any message was sent
        current_status = supplier.get("outreach_status", "none")
        new_status = current_status
        if email_sent or wechat_sent:
            new_status = "contacted"
            supplier_repository.update_outreach_status(supplier_id, new_status)

        print(
            f"INFO [SupplierService]: Outreach sent for supplier {supplier_id} "
            f"(email={email_sent}, wechat={wechat_sent}, mock={mock_mode}, status={new_status})"
        )

        return {
            "email_sent": email_sent,
            "wechat_sent": wechat_sent,
            "message": "; ".join(messages),
            "outreach_status": new_status,
            "mock_mode": mock_mode,
        }

    def _format_markets_served(self, markets: Any) -> str:
        """Format markets_served JSONB dict to readable string.

        Args:
            markets: Dict like {"Asia": 60, "Europe": 30} or None

        Returns:
            Formatted string like "Asia: 60%, Europe: 30%"
        """
        if not markets or not isinstance(markets, dict):
            return ""
        return ", ".join(f"{k}: {v}%" for k, v in markets.items())

    def _format_list(self, items: Any, separator: str = ", ") -> str:
        """Format a list/array field to a joined string.

        Args:
            items: List of strings or None
            separator: Separator to join with

        Returns:
            Joined string or empty string
        """
        if not items or not isinstance(items, list):
            return ""
        return separator.join(str(item) for item in items)

    def export_verification_excel(
        self,
        status: Optional[str] = None,
        country: Optional[str] = None,
        has_products: Optional[bool] = None,
        certification_status: Optional[str] = None,
        pipeline_status: Optional[str] = None,
        search: Optional[str] = None,
        sort_by: str = "name",
        sort_order: str = "asc",
    ) -> bytes:
        """Export supplier verification data to Excel.

        Args:
            status: Filter by supplier status
            country: Filter by country
            has_products: Filter by whether supplier has products
            certification_status: Filter by certification status
            pipeline_status: Filter by pipeline status
            search: Search query
            sort_by: Field to sort by
            sort_order: Sort direction

        Returns:
            Excel file bytes
        """
        items = supplier_repository.get_all_with_audit_data(
            status=status,
            country=country,
            has_products=has_products,
            certification_status=certification_status,
            pipeline_status=pipeline_status,
            search=search,
            sort_by=sort_by,
            sort_order=sort_order,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Supplier Verification Data"

        # Define columns
        columns = [
            # Supplier Info
            ("Name", "name"),
            ("Code", "code"),
            ("Status", "status"),
            ("Country", "country"),
            ("City", "city"),
            ("Contact Name", "contact_name"),
            ("Contact Email", "contact_email"),
            ("Contact Phone", "contact_phone"),
            ("Website", "website"),
            # Certification Info
            ("Certification Status", "certification_status"),
            ("Pipeline Status", "pipeline_status"),
            ("Certified At", "certified_at"),
            # Audit Info
            ("Audit Date", "audit_date"),
            ("Inspector Name", "inspector_name"),
            ("Extraction Status", "extraction_status"),
            # Extracted Data
            ("Supplier Type", "supplier_type"),
            ("Employee Count", "employee_count"),
            ("Factory Area (sqm)", "factory_area_sqm"),
            ("Production Lines", "production_lines_count"),
            ("Certifications", "certifications"),
            ("Markets Served", "markets_served"),
            ("Has Machinery Photos", "has_machinery_photos"),
            ("Positive Points", "positive_points"),
            ("Negative Points", "negative_points"),
            ("Products Verified", "products_verified"),
            # Classification
            ("AI Classification", "ai_classification"),
            ("AI Classification Reason", "ai_classification_reason"),
            ("Manual Classification", "manual_classification"),
            ("Classification Notes", "classification_notes"),
        ]

        # Write header row with formatting
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        for col_idx, (header, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data rows
        for row_idx, item in enumerate(items, 2):
            for col_idx, (_, field) in enumerate(columns, 1):
                value = item.get(field)

                # Format special fields
                if field == "certifications":
                    value = self._format_list(value, ", ")
                elif field == "markets_served":
                    value = self._format_markets_served(value)
                elif field == "positive_points":
                    value = self._format_list(value, "; ")
                elif field == "negative_points":
                    value = self._format_list(value, "; ")
                elif field == "products_verified":
                    value = self._format_list(value, ", ")
                elif field == "has_machinery_photos":
                    value = "Yes" if value else ("No" if value is False else "")
                elif value is None:
                    value = ""

                ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else value)

        # Auto-size columns (approximate)
        for col_idx, (header, _) in enumerate(columns, 1):
            max_length = len(header)
            for row_idx in range(2, min(len(items) + 2, 52)):  # Sample first 50 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 50)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        print(f"INFO [SupplierService]: Exported {len(items)} suppliers to Excel")
        return buffer.getvalue()


# Singleton instance
supplier_service = SupplierService()
