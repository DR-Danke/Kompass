"""AI-powered data extraction service for product catalogs."""

import base64
import io
import json
import time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional, Tuple

import httpx
from PIL import Image

from app.config import get_settings
from app.models.extraction_dto import (
    ExtractedProduct,
    ExtractionResult,
    HsCodeSuggestion,
    ImageOperation,
    ImageProcessingResult,
)

# File size limits
MAX_PDF_SIZE_MB = 50
MAX_IMAGE_SIZE_MB = 10
MAX_EXCEL_SIZE_MB = 20


class ExtractionService:
    """Service for extracting product data from various document formats."""

    def __init__(self) -> None:
        """Initialize the extraction service with AI clients."""
        self._anthropic_client = None
        self._openai_client = None
        self._settings = get_settings()

    def _get_anthropic_client(self):
        """Lazily initialize and return the Anthropic client."""
        if self._anthropic_client is None and self._settings.ANTHROPIC_API_KEY:
            try:
                import anthropic

                self._anthropic_client = anthropic.Anthropic(
                    api_key=self._settings.ANTHROPIC_API_KEY
                )
            except ImportError:
                print("WARN [ExtractionService]: anthropic package not installed")
        return self._anthropic_client

    def _get_openai_client(self):
        """Lazily initialize and return the OpenAI client."""
        if self._openai_client is None and self._settings.OPENAI_API_KEY:
            try:
                import openai

                self._openai_client = openai.OpenAI(
                    api_key=self._settings.OPENAI_API_KEY
                )
            except ImportError:
                print("WARN [ExtractionService]: openai package not installed")
        return self._openai_client

    def _is_ai_available(self) -> bool:
        """Check if any AI service is configured and available."""
        return bool(
            self._settings.ANTHROPIC_API_KEY or self._settings.OPENAI_API_KEY
        )

    def _get_preferred_ai_provider(self) -> str:
        """Get the preferred AI provider based on configuration."""
        provider = self._settings.EXTRACTION_AI_PROVIDER.lower()
        if provider == "anthropic" and self._settings.ANTHROPIC_API_KEY:
            return "anthropic"
        if provider == "openai" and self._settings.OPENAI_API_KEY:
            return "openai"
        # Fallback to whatever is available
        if self._settings.ANTHROPIC_API_KEY:
            return "anthropic"
        if self._settings.OPENAI_API_KEY:
            return "openai"
        return "none"

    def _build_extraction_prompt(self) -> str:
        """Build the prompt for product data extraction."""
        return """Extract product information from the provided content.
Return a JSON object with these fields (use null for missing values):
{
    "sku": "product SKU/reference code",
    "name": "product name",
    "description": "product description",
    "price_fob_usd": decimal price in USD,
    "moq": minimum order quantity as integer,
    "dimensions": "dimensions string (e.g., 10x20x30cm)",
    "material": "material description",
    "suggested_category": "suggested product category"
}

Only return valid JSON, no additional text or explanation."""

    def _parse_extraction_response(
        self, response_text: str, source_page: Optional[int] = None
    ) -> ExtractedProduct:
        """Parse AI response into ExtractedProduct DTO."""
        try:
            # Try to extract JSON from response
            json_start = response_text.find("{")
            json_end = response_text.rfind("}") + 1
            if json_start >= 0 and json_end > json_start:
                json_str = response_text[json_start:json_end]
                data = json.loads(json_str)
            else:
                data = {}
        except json.JSONDecodeError:
            print(f"WARN [ExtractionService]: Failed to parse JSON: {response_text}")
            data = {}

        # Calculate confidence based on fields extracted
        fields_present = sum(
            1 for k in ["sku", "name", "description", "price_fob_usd", "moq"]
            if data.get(k) is not None
        )
        confidence = fields_present / 5.0

        # Parse price safely
        price_fob_usd = None
        if data.get("price_fob_usd") is not None:
            try:
                price_fob_usd = Decimal(str(data["price_fob_usd"]))
            except (InvalidOperation, ValueError):
                pass

        return ExtractedProduct(
            sku=data.get("sku"),
            name=data.get("name"),
            description=data.get("description"),
            price_fob_usd=price_fob_usd,
            moq=data.get("moq"),
            dimensions=data.get("dimensions"),
            material=data.get("material"),
            suggested_category=data.get("suggested_category"),
            confidence_score=confidence,
            source_page=source_page,
        )

    def _extract_with_anthropic(
        self, content: str, is_image: bool = False, image_data: Optional[bytes] = None
    ) -> str:
        """Extract product data using Claude API."""
        client = self._get_anthropic_client()
        if not client:
            raise RuntimeError("Anthropic client not available")

        prompt = self._build_extraction_prompt()

        if is_image and image_data:
            # Use vision capability
            base64_image = base64.b64encode(image_data).decode("utf-8")
            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1024,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/jpeg",
                                    "data": base64_image,
                                },
                            },
                            {"type": "text", "text": prompt},
                        ],
                    }
                ],
            )
        else:
            # Text extraction
            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1024,
                messages=[
                    {
                        "role": "user",
                        "content": f"{prompt}\n\nContent to extract from:\n{content}",
                    }
                ],
            )

        return message.content[0].text

    def _extract_with_openai(
        self, content: str, is_image: bool = False, image_data: Optional[bytes] = None
    ) -> str:
        """Extract product data using OpenAI API."""
        client = self._get_openai_client()
        if not client:
            raise RuntimeError("OpenAI client not available")

        prompt = self._build_extraction_prompt()

        if is_image and image_data:
            # Use vision capability
            base64_image = base64.b64encode(image_data).decode("utf-8")
            response = client.chat.completions.create(
                model="gpt-4o",
                max_tokens=1024,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}"
                                },
                            },
                        ],
                    }
                ],
            )
        else:
            # Text extraction
            response = client.chat.completions.create(
                model="gpt-4o",
                max_tokens=1024,
                messages=[
                    {
                        "role": "user",
                        "content": f"{prompt}\n\nContent to extract from:\n{content}",
                    }
                ],
            )

        return response.choices[0].message.content

    def extract_product_data(
        self,
        content: str,
        is_image: bool = False,
        image_data: Optional[bytes] = None,
        source_page: Optional[int] = None,
    ) -> ExtractedProduct:
        """Extract product data from text or image using AI.

        Args:
            content: Text content to extract from (or description for images)
            is_image: Whether image_data contains an image
            image_data: Raw image bytes if is_image is True
            source_page: Optional source page number

        Returns:
            ExtractedProduct with extracted fields and confidence score
        """
        if not self._is_ai_available():
            print("INFO [ExtractionService]: AI unavailable, returning empty product")
            return ExtractedProduct(
                raw_text=content if not is_image else None,
                source_page=source_page,
                confidence_score=0.0,
            )

        provider = self._get_preferred_ai_provider()

        try:
            if provider == "anthropic":
                response = self._extract_with_anthropic(content, is_image, image_data)
            else:
                response = self._extract_with_openai(content, is_image, image_data)

            product = self._parse_extraction_response(response, source_page)
            product.raw_text = content if not is_image else None
            return product

        except Exception as e:
            print(f"WARN [ExtractionService]: AI extraction failed: {e}")
            return ExtractedProduct(
                raw_text=content if not is_image else None,
                source_page=source_page,
                confidence_score=0.0,
            )

    def process_pdf(self, file_path: str) -> Tuple[List[ExtractedProduct], List[str]]:
        """Process a PDF file and extract product data.

        Args:
            file_path: Path to the PDF file

        Returns:
            Tuple of (extracted products, errors)
        """
        from PyPDF2 import PdfReader

        products: List[ExtractedProduct] = []
        errors: List[str] = []

        path = Path(file_path)
        if not path.exists():
            return [], [f"File not found: {file_path}"]

        # Check file size
        file_size_mb = path.stat().st_size / (1024 * 1024)
        if file_size_mb > MAX_PDF_SIZE_MB:
            return [], [f"PDF exceeds {MAX_PDF_SIZE_MB}MB limit"]

        try:
            reader = PdfReader(file_path)

            if reader.is_encrypted:
                return [], ["PDF is encrypted and cannot be processed"]

            for page_num, page in enumerate(reader.pages, start=1):
                try:
                    text = page.extract_text() or ""

                    if len(text.strip()) < 50:
                        # Not enough text, skip this page
                        continue

                    product = self.extract_product_data(
                        content=text, is_image=False, source_page=page_num
                    )
                    if product.name or product.sku:
                        products.append(product)

                except Exception as e:
                    errors.append(f"Error processing page {page_num}: {e}")

        except Exception as e:
            errors.append(f"Error reading PDF: {e}")

        return products, errors

    # Column name mappings (lowercase variations) including Spanish and format variants
    _sku_columns = [
        "sku", "reference", "code", "ref", "item code", "product code",
        "referencia", "ref.", "modelo", "item no", "item no.", "no.",
    ]
    _name_columns = [
        "name", "product", "item", "product name", "item name", "title",
        "producto", "descripcion producto", "nombre", "product type",
    ]
    _price_columns = [
        "price", "fob", "fob price", "unit price", "cost", "price usd",
        "precio", "fob u/p", "u/p", "unit price(usd)", "price (usd/m2)",
        "costo unitario",
    ]
    _moq_columns = [
        "moq", "minimum", "min qty", "min order", "minimum order",
        "qty", "quantity", "cantidad",
    ]
    _description_columns = [
        "description", "desc", "details", "product description",
        "caracteristics", "características", "caracteristicas",
    ]
    _material_columns = [
        "material", "materials", "composition",
        "finish", "acabado", "surface", "collection",
    ]
    _dimensions_columns = [
        "dimensions", "size", "sizes", "dim",
        "size(mm)", "specification", "medida",
    ]

    # Unit-of-measure keywords to detect from price column headers
    _unit_mapping = {
        "m2": "m2", "sqm": "m2", "m²": "m2",
        "pcs": "piece", "pc": "piece", "piece": "piece",
        "set": "set", "pair": "pair",
        "kg": "kg", "ton": "ton",
        "meter": "meter", "m ": "meter",
    }

    @staticmethod
    def _find_column(
        headers: List[str], candidates: List[str]
    ) -> Optional[int]:
        """Find column index using exact-then-substring matching.

        First tries exact match (lowercased). If no exact match found,
        tries substring matching where any candidate is contained in
        the lowercased header text.
        """
        # Pass 1: exact match
        for idx, header in enumerate(headers):
            if not header:
                continue
            header_lower = header.lower().strip()
            if not header_lower:
                continue
            if header_lower in candidates:
                return idx
        # Pass 2: substring (candidate contained in header)
        for idx, header in enumerate(headers):
            if not header:
                continue
            header_lower = header.lower().strip()
            if not header_lower:
                continue
            for candidate in candidates:
                if candidate in header_lower:
                    return idx
        return None

    def _find_best_header_row(
        self, rows: list
    ) -> Tuple[int, dict, int]:
        """Scan up to the first 10 rows to find the best header row.

        Returns:
            Tuple of (header_row_index, column_mapping dict, score)
            where column_mapping maps category name to column index.
        """
        # Ordered so that more specific categories are matched first,
        # preventing generic candidates (e.g. "item") from stealing columns.
        all_categories = [
            ("sku", self._sku_columns),
            ("price", self._price_columns),
            ("moq", self._moq_columns),
            ("description", self._description_columns),
            ("material", self._material_columns),
            ("dimensions", self._dimensions_columns),
            ("name", self._name_columns),
        ]

        best_row_idx = 0
        best_mapping: dict = {}
        best_score = 0

        scan_limit = min(10, len(rows))
        for row_idx in range(scan_limit):
            headers = [str(h) if h else "" for h in rows[row_idx]]
            mapping: dict = {}
            claimed_cols: set = set()
            score = 0
            for cat_name, candidates in all_categories:
                col_idx = self._find_column(headers, candidates)
                if col_idx is not None and col_idx not in claimed_cols:
                    mapping[cat_name] = col_idx
                    claimed_cols.add(col_idx)
                    score += 1

            if score > best_score:
                best_score = score
                best_row_idx = row_idx
                best_mapping = mapping

        return best_row_idx, best_mapping, best_score

    def _detect_unit_of_measure(self, headers: List[str], price_idx: Optional[int]) -> Optional[str]:
        """Detect unit of measure from the price column header text."""
        if price_idx is None or price_idx >= len(headers):
            return None
        header_text = headers[price_idx].lower() if headers[price_idx] else ""
        for keyword, unit in self._unit_mapping.items():
            if keyword in header_text:
                return unit
        return None

    def _extract_excel_with_ai(
        self, rows: list, sheet_title: str
    ) -> List[ExtractedProduct]:
        """Use AI to extract products from unrecognized Excel data.

        Serializes the first 50 data rows as a text table and sends
        to the preferred AI provider for structured extraction.
        """
        # Serialize first 50 rows as pipe-separated table
        data_rows = rows[:50]
        lines: List[str] = []
        for row in data_rows:
            cells = [str(c) if c is not None else "" for c in row]
            lines.append(" | ".join(cells))
        table_text = "\n".join(lines)

        prompt = (
            "Extract product data from this spreadsheet table. Each row may represent a product.\n"
            "Return a JSON array of objects, each with: sku, name, description, "
            "price_fob_usd (decimal), moq (integer), dimensions, material.\n"
            "Use null for missing values. Only return valid JSON array, no additional text.\n\n"
            f"Table data:\n{table_text}"
        )

        provider = self._get_preferred_ai_provider()
        try:
            if provider == "anthropic":
                client = self._get_anthropic_client()
                if not client:
                    raise RuntimeError("Anthropic client not available")
                message = client.messages.create(
                    model="claude-sonnet-4-20250514",
                    max_tokens=4096,
                    messages=[{"role": "user", "content": prompt}],
                )
                response_text = message.content[0].text
            elif provider == "openai":
                client = self._get_openai_client()
                if not client:
                    raise RuntimeError("OpenAI client not available")
                response = client.chat.completions.create(
                    model="gpt-4o",
                    max_tokens=4096,
                    messages=[{"role": "user", "content": prompt}],
                )
                response_text = response.choices[0].message.content
            else:
                return []

            # Parse JSON array response
            json_start = response_text.find("[")
            json_end = response_text.rfind("]") + 1
            if json_start >= 0 and json_end > json_start:
                items = json.loads(response_text[json_start:json_end])
            else:
                # Try single object fallback
                product = self._parse_extraction_response(response_text)
                return [product] if product.name or product.sku else []

            products: List[ExtractedProduct] = []
            for item in items:
                if not isinstance(item, dict):
                    continue
                # Calculate confidence
                fields_present = sum(
                    1 for k in ["sku", "name", "description", "price_fob_usd", "moq"]
                    if item.get(k) is not None
                )
                confidence = fields_present / 5.0

                price_fob_usd = None
                if item.get("price_fob_usd") is not None:
                    try:
                        price_fob_usd = Decimal(str(item["price_fob_usd"]))
                    except (InvalidOperation, ValueError):
                        pass

                product = ExtractedProduct(
                    sku=item.get("sku"),
                    name=item.get("name"),
                    description=item.get("description"),
                    price_fob_usd=price_fob_usd,
                    moq=item.get("moq"),
                    dimensions=item.get("dimensions"),
                    material=item.get("material"),
                    confidence_score=confidence,
                )
                if product.name or product.sku:
                    products.append(product)
            return products

        except Exception as e:
            print(f"WARN [ExtractionService]: AI Excel extraction failed for sheet '{sheet_title}': {e}")
            return []

    def process_excel(
        self, file_path: str
    ) -> Tuple[List[ExtractedProduct], List[str]]:
        """Process an Excel file and extract product data.

        Uses multi-row header scanning (first 10 rows), substring matching
        for column detection, unit-of-measure detection from price headers,
        and AI fallback for unrecognized formats.

        Args:
            file_path: Path to the Excel file

        Returns:
            Tuple of (extracted products, errors)
        """
        from openpyxl import load_workbook

        products: List[ExtractedProduct] = []
        errors: List[str] = []

        path = Path(file_path)
        if not path.exists():
            return [], [f"File not found: {file_path}"]

        # Check file size
        file_size_mb = path.stat().st_size / (1024 * 1024)
        if file_size_mb > MAX_EXCEL_SIZE_MB:
            return [], [f"Excel exceeds {MAX_EXCEL_SIZE_MB}MB limit"]

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)

            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue

                # Find best header row by scanning first 10 rows
                header_row_idx, col_mapping, score = self._find_best_header_row(rows)

                # AI fallback when fewer than 2 column categories match
                if score < 2:
                    print(
                        f"WARN [ExtractionService]: AI fallback used for sheet "
                        f"'{sheet.title}' - no header matches found"
                    )
                    if not self._is_ai_available():
                        print(
                            f"WARN [ExtractionService]: AI unavailable, "
                            f"skipping sheet '{sheet.title}'"
                        )
                        continue
                    ai_products = self._extract_excel_with_ai(rows, sheet.title)
                    products.extend(ai_products)
                    continue

                # Use the detected header row
                headers = [str(h) if h else "" for h in rows[header_row_idx]]
                sku_idx = col_mapping.get("sku")
                name_idx = col_mapping.get("name")
                price_idx = col_mapping.get("price")
                moq_idx = col_mapping.get("moq")
                desc_idx = col_mapping.get("description")
                material_idx = col_mapping.get("material")
                dim_idx = col_mapping.get("dimensions")

                # Fallback: use description column as name when no name column found
                if name_idx is None and desc_idx is not None:
                    name_idx = desc_idx
                    desc_idx = None

                # Detect unit of measure from price column header
                unit_of_measure = self._detect_unit_of_measure(headers, price_idx)

                # Process data rows starting after header row
                for row_num, row in enumerate(
                    rows[header_row_idx + 1:], start=header_row_idx + 2
                ):
                    try:
                        # Skip empty rows
                        if not any(row):
                            continue

                        def get_cell(idx: Optional[int]) -> Optional[str]:
                            if idx is None or idx >= len(row) or row[idx] is None:
                                return None
                            return str(row[idx]).strip() or None

                        sku = get_cell(sku_idx)
                        name = get_cell(name_idx)

                        # Skip rows without name or SKU
                        if not name and not sku:
                            continue

                        # Parse price
                        price_fob_usd = None
                        price_str = get_cell(price_idx)
                        if price_str:
                            try:
                                # Remove currency symbols
                                price_clean = price_str.replace("$", "").replace(",", "")
                                price_fob_usd = Decimal(price_clean)
                            except (InvalidOperation, ValueError):
                                pass

                        # Parse MOQ
                        moq = None
                        moq_str = get_cell(moq_idx)
                        if moq_str:
                            try:
                                moq = int(float(moq_str))
                            except (ValueError, TypeError):
                                pass

                        product = ExtractedProduct(
                            sku=sku,
                            name=name,
                            description=get_cell(desc_idx),
                            price_fob_usd=price_fob_usd,
                            moq=moq,
                            material=get_cell(material_idx),
                            dimensions=get_cell(dim_idx),
                            confidence_score=0.8,  # High confidence for structured data
                            unit_of_measure=unit_of_measure,
                        )
                        products.append(product)

                    except Exception as e:
                        errors.append(f"Error processing row {row_num}: {e}")

            workbook.close()

        except Exception as e:
            errors.append(f"Error reading Excel: {e}")

        return products, errors

    def process_image(
        self, file_path: str
    ) -> Tuple[Optional[ExtractedProduct], List[str]]:
        """Process a single image file and extract product data.

        Args:
            file_path: Path to the image file

        Returns:
            Tuple of (extracted product or None, errors)
        """
        errors: List[str] = []

        path = Path(file_path)
        if not path.exists():
            return None, [f"File not found: {file_path}"]

        # Check file size
        file_size_mb = path.stat().st_size / (1024 * 1024)
        if file_size_mb > MAX_IMAGE_SIZE_MB:
            return None, [f"Image exceeds {MAX_IMAGE_SIZE_MB}MB limit"]

        try:
            with open(file_path, "rb") as f:
                image_data = f.read()

            product = self.extract_product_data(
                content="Extract product information from this image",
                is_image=True,
                image_data=image_data,
            )
            return product, []

        except Exception as e:
            errors.append(f"Error processing image: {e}")
            return None, errors

    def process_batch(self, file_paths: List[str]) -> ExtractionResult:
        """Process multiple files and aggregate results.

        Args:
            file_paths: List of file paths to process

        Returns:
            ExtractionResult with all extracted products and errors
        """
        start_time = time.time()
        all_products: List[ExtractedProduct] = []
        all_errors: List[str] = []
        all_warnings: List[str] = []

        for file_path in file_paths:
            path = Path(file_path)
            suffix = path.suffix.lower()

            try:
                if suffix == ".pdf":
                    products, errors = self.process_pdf(file_path)
                    all_products.extend(products)
                    all_errors.extend(errors)

                elif suffix in [".xlsx", ".xls"]:
                    products, errors = self.process_excel(file_path)
                    all_products.extend(products)
                    all_errors.extend(errors)

                elif suffix in [".jpg", ".jpeg", ".png", ".gif", ".webp"]:
                    product, errors = self.process_image(file_path)
                    if product:
                        all_products.append(product)
                    all_errors.extend(errors)

                else:
                    all_warnings.append(f"Unsupported file type: {suffix}")

            except Exception as e:
                all_errors.append(f"Error processing {file_path}: {e}")

        processing_time = time.time() - start_time

        if not self._is_ai_available():
            all_warnings.append(
                "AI services unavailable - extraction quality may be reduced"
            )

        return ExtractionResult(
            products=all_products,
            total_extracted=len(all_products),
            errors=all_errors,
            warnings=all_warnings,
            processing_time_seconds=round(processing_time, 2),
        )

    def suggest_hs_code(self, product_description: str) -> HsCodeSuggestion:
        """Suggest an HS code based on product description.

        Args:
            product_description: Description of the product

        Returns:
            HsCodeSuggestion with suggested code and confidence
        """
        if not self._is_ai_available():
            print("INFO [ExtractionService]: AI unavailable for HS code suggestion")
            return HsCodeSuggestion(
                code="9999.99.99",
                description="Unable to classify - AI service unavailable",
                confidence_score=0.0,
                reasoning="AI service not configured",
            )

        prompt = f"""Suggest the most appropriate HS (Harmonized System) code for this product.
Consider common categories for China imports to Latin America.

Product description: {product_description}

Return a JSON object with:
{{
    "code": "HS code in format XXXX.XX.XX",
    "description": "Official HS code description",
    "confidence_score": 0.0 to 1.0,
    "reasoning": "Brief explanation of why this code was selected"
}}

Only return valid JSON, no additional text."""

        try:
            provider = self._get_preferred_ai_provider()

            if provider == "anthropic":
                client = self._get_anthropic_client()
                message = client.messages.create(
                    model="claude-sonnet-4-20250514",
                    max_tokens=512,
                    messages=[{"role": "user", "content": prompt}],
                )
                response_text = message.content[0].text
            else:
                client = self._get_openai_client()
                response = client.chat.completions.create(
                    model="gpt-4o",
                    max_tokens=512,
                    messages=[{"role": "user", "content": prompt}],
                )
                response_text = response.choices[0].message.content

            # Parse response
            json_start = response_text.find("{")
            json_end = response_text.rfind("}") + 1
            if json_start >= 0 and json_end > json_start:
                data = json.loads(response_text[json_start:json_end])
                return HsCodeSuggestion(
                    code=data.get("code", "9999.99.99"),
                    description=data.get("description", "Unknown"),
                    confidence_score=float(data.get("confidence_score", 0.5)),
                    reasoning=data.get("reasoning"),
                )

        except Exception as e:
            print(f"WARN [ExtractionService]: HS code suggestion failed: {e}")

        return HsCodeSuggestion(
            code="9999.99.99",
            description="Classification failed",
            confidence_score=0.0,
            reasoning="Error during classification",
        )

    def remove_background(self, image_url: str) -> ImageProcessingResult:
        """Remove background from an image using RemoveBG API.

        Args:
            image_url: URL of the image to process

        Returns:
            ImageProcessingResult with original and processed URLs
        """
        if not self._settings.REMOVEBG_API_KEY:
            print("INFO [ExtractionService]: RemoveBG API key not configured")
            return ImageProcessingResult(
                original_url=image_url,
                processed_url=image_url,
                operation=ImageOperation.REMOVE_BG,
            )

        try:
            with httpx.Client(timeout=self._settings.EXTRACTION_TIMEOUT_SECONDS) as client:
                response = client.post(
                    "https://api.remove.bg/v1.0/removebg",
                    headers={"X-Api-Key": self._settings.REMOVEBG_API_KEY},
                    data={"image_url": image_url, "size": "auto"},
                )

                if response.status_code == 200:
                    # Return as data URI
                    base64_image = base64.b64encode(response.content).decode("utf-8")
                    processed_url = f"data:image/png;base64,{base64_image}"
                    return ImageProcessingResult(
                        original_url=image_url,
                        processed_url=processed_url,
                        operation=ImageOperation.REMOVE_BG,
                    )
                else:
                    print(
                        f"WARN [ExtractionService]: RemoveBG failed: {response.status_code}"
                    )

        except Exception as e:
            print(f"WARN [ExtractionService]: RemoveBG error: {e}")

        return ImageProcessingResult(
            original_url=image_url,
            processed_url=image_url,
            operation=ImageOperation.REMOVE_BG,
        )

    def resize_image(
        self, image_path: str, width: int, height: int
    ) -> Tuple[Optional[str], Optional[str]]:
        """Resize an image while maintaining aspect ratio.

        Args:
            image_path: Path to the image file
            width: Target width
            height: Target height

        Returns:
            Tuple of (data URI of resized image, error message if any)
        """
        try:
            with Image.open(image_path) as img:
                # Maintain aspect ratio
                img.thumbnail((width, height), Image.Resampling.LANCZOS)

                # Convert to bytes
                buffer = io.BytesIO()
                img_format = img.format or "PNG"
                img.save(buffer, format=img_format)
                buffer.seek(0)

                # Return as data URI
                base64_image = base64.b64encode(buffer.read()).decode("utf-8")
                mime_type = f"image/{img_format.lower()}"
                return f"data:{mime_type};base64,{base64_image}", None

        except Exception as e:
            return None, f"Error resizing image: {e}"

    def find_higher_quality_image(self, image_url: str) -> Optional[str]:
        """Find a higher quality version of an image.

        Note: This is a placeholder implementation. Reverse image search
        integration (Google, TinEye) would require additional API setup.

        Args:
            image_url: URL of the original image

        Returns:
            URL of higher quality image if found, None otherwise
        """
        # TODO: Implement reverse image search integration
        # Options: Google Vision API, TinEye API, or custom solution
        print("INFO [ExtractionService]: Higher quality image search not implemented")
        return None


# Singleton instance
extraction_service = ExtractionService()
