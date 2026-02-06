"""Unit tests for supplier API routes."""

from datetime import datetime
from decimal import Decimal
from unittest.mock import patch
from uuid import uuid4

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.api.dependencies import get_current_user
from app.api.supplier_routes import router
from app.models.kompass_dto import (
    PaginationDTO,
    ProductFilterDTO,
    ProductListResponseDTO,
    ProductResponseDTO,
    ProductStatus,
    SupplierListResponseDTO,
    SupplierResponseDTO,
    SupplierStatus,
)


@pytest.fixture
def mock_current_user():
    """Mock current user for authentication."""
    return {
        "id": str(uuid4()),
        "email": "test@example.com",
        "role": "user",
        "is_active": True,
    }


@pytest.fixture
def mock_admin_user():
    """Mock admin user for RBAC tests."""
    return {
        "id": str(uuid4()),
        "email": "admin@example.com",
        "role": "admin",
        "is_active": True,
    }


@pytest.fixture
def sample_supplier_data():
    """Sample supplier data for testing."""
    return {
        "id": uuid4(),
        "name": "Test Supplier",
        "code": "TS001",
        "status": SupplierStatus.ACTIVE,
        "contact_name": "John Doe",
        "contact_email": "john@example.com",
        "contact_phone": "123456789",
        "address": "123 Test Street",
        "city": "Shanghai",
        "country": "China",
        "website": "https://example.com",
        "notes": "Test notes",
        "created_at": datetime.now(),
        "updated_at": datetime.now(),
    }


@pytest.fixture
def sample_supplier_response(sample_supplier_data):
    """Create SupplierResponseDTO from sample data."""
    return SupplierResponseDTO(**sample_supplier_data)


@pytest.fixture
def sample_product_data():
    """Sample product data for testing."""
    return {
        "id": uuid4(),
        "sku": "PRD-001",
        "name": "Test Product",
        "description": "Test description",
        "supplier_id": uuid4(),
        "supplier_name": "Test Supplier",
        "category_id": None,
        "category_name": None,
        "hs_code_id": None,
        "hs_code": None,
        "status": ProductStatus.ACTIVE,
        "unit_cost": Decimal("10.00"),
        "unit_price": Decimal("20.00"),
        "currency": "USD",
        "unit_of_measure": "piece",
        "minimum_order_qty": 1,
        "lead_time_days": None,
        "weight_kg": None,
        "dimensions": None,
        "origin_country": "China",
        "images": [],
        "tags": [],
        "created_at": datetime.now(),
        "updated_at": datetime.now(),
    }


@pytest.fixture
def app(mock_current_user):
    """Create test FastAPI app with dependency overrides."""
    test_app = FastAPI()
    test_app.include_router(router, prefix="/api/suppliers")

    # Override auth dependency
    test_app.dependency_overrides[get_current_user] = lambda: mock_current_user

    return test_app


@pytest.fixture
def app_admin(mock_admin_user):
    """Create test FastAPI app with admin user."""
    test_app = FastAPI()
    test_app.include_router(router, prefix="/api/suppliers")

    # Override auth dependency
    test_app.dependency_overrides[get_current_user] = lambda: mock_admin_user

    # Override RBAC dependency for admin
    def override_require_roles(allowed_roles):
        async def role_checker():
            return mock_admin_user
        return role_checker

    # Note: We need a different approach for require_roles since it's a factory
    return test_app


@pytest.fixture
def client(app):
    """Create test client."""
    return TestClient(app)


@pytest.fixture
def admin_client(mock_admin_user):
    """Create test client with admin user for RBAC tests."""
    test_app = FastAPI()
    test_app.include_router(router, prefix="/api/suppliers")

    # Override auth dependency
    test_app.dependency_overrides[get_current_user] = lambda: mock_admin_user

    return TestClient(test_app)


class TestListSuppliers:
    """Tests for GET /suppliers endpoint."""

    @patch("app.api.supplier_routes.supplier_service")
    def test_list_suppliers_success(
        self, mock_service, client, sample_supplier_response
    ):
        """Test successful supplier listing."""
        mock_service.list_suppliers.return_value = SupplierListResponseDTO(
            items=[sample_supplier_response],
            pagination=PaginationDTO(page=1, limit=20, total=1, pages=1),
        )

        response = client.get("/api/suppliers")

        assert response.status_code == 200
        data = response.json()
        assert len(data["items"]) == 1
        assert data["pagination"]["page"] == 1

    @patch("app.api.supplier_routes.supplier_service")
    def test_list_suppliers_with_filters(
        self, mock_service, client, sample_supplier_response
    ):
        """Test supplier listing with filters."""
        mock_service.list_suppliers.return_value = SupplierListResponseDTO(
            items=[sample_supplier_response],
            pagination=PaginationDTO(page=1, limit=20, total=1, pages=1),
        )

        response = client.get(
            "/api/suppliers?status=active&country=China&has_products=true&page=2&limit=10"
        )

        assert response.status_code == 200
        mock_service.list_suppliers.assert_called_once()
        call_kwargs = mock_service.list_suppliers.call_args.kwargs
        assert call_kwargs["status"] == SupplierStatus.ACTIVE
        assert call_kwargs["country"] == "China"
        assert call_kwargs["has_products"] is True
        assert call_kwargs["page"] == 2
        assert call_kwargs["limit"] == 10

    @patch("app.api.supplier_routes.supplier_service")
    def test_list_suppliers_invalid_status(self, mock_service, client):
        """Test supplier listing with invalid status filter."""
        response = client.get("/api/suppliers?status=invalid_status")

        assert response.status_code == 400
        assert "Invalid status value" in response.json()["detail"]

    def test_list_suppliers_unauthenticated(self):
        """Test that unauthenticated requests are rejected."""
        test_app = FastAPI()
        test_app.include_router(router, prefix="/api/suppliers")
        # No dependency override - uses real auth
        unauth_client = TestClient(test_app)
        response = unauth_client.get("/api/suppliers")

        assert response.status_code in [401, 403]


class TestCreateSupplier:
    """Tests for POST /suppliers endpoint."""

    @patch("app.api.supplier_routes.supplier_service")
    def test_create_supplier_success(
        self, mock_service, client, sample_supplier_response
    ):
        """Test successful supplier creation."""
        mock_service.create_supplier.return_value = sample_supplier_response

        response = client.post(
            "/api/suppliers",
            json={"name": "New Supplier", "country": "China"},
        )

        assert response.status_code == 201
        assert response.json()["name"] == "Test Supplier"
        mock_service.create_supplier.assert_called_once()

    @patch("app.api.supplier_routes.supplier_service")
    def test_create_supplier_validation_error(self, mock_service, client):
        """Test supplier creation with validation error."""
        mock_service.create_supplier.side_effect = ValueError("Invalid email format")

        response = client.post(
            "/api/suppliers",
            json={"name": "New Supplier"},
        )

        assert response.status_code == 400
        assert "Invalid email format" in response.json()["detail"]

    def test_create_supplier_unauthenticated(self):
        """Test that unauthenticated creation is rejected."""
        test_app = FastAPI()
        test_app.include_router(router, prefix="/api/suppliers")
        unauth_client = TestClient(test_app)
        response = unauth_client.post("/api/suppliers", json={"name": "Test"})

        assert response.status_code in [401, 403]


class TestSearchSuppliers:
    """Tests for GET /suppliers/search endpoint."""

    @patch("app.api.supplier_routes.supplier_service")
    def test_search_suppliers_success(
        self, mock_service, client, sample_supplier_response
    ):
        """Test successful supplier search."""
        mock_service.search_suppliers.return_value = [sample_supplier_response]

        response = client.get("/api/suppliers/search?query=test")

        assert response.status_code == 200
        data = response.json()
        assert len(data) == 1
        mock_service.search_suppliers.assert_called_once_with("test")

    @patch("app.api.supplier_routes.supplier_service")
    def test_search_suppliers_empty_results(self, mock_service, client):
        """Test search with no results."""
        mock_service.search_suppliers.return_value = []

        response = client.get("/api/suppliers/search?query=nonexistent")

        assert response.status_code == 200
        assert response.json() == []

    def test_search_suppliers_short_query(self, client):
        """Test search with query too short."""
        response = client.get("/api/suppliers/search?query=a")

        assert response.status_code == 422  # Validation error


class TestGetSupplier:
    """Tests for GET /suppliers/{supplier_id} endpoint."""

    @patch("app.api.supplier_routes.supplier_service")
    def test_get_supplier_success(
        self, mock_service, client, sample_supplier_data
    ):
        """Test successful supplier retrieval."""
        supplier_dict = dict(sample_supplier_data)
        # Convert status enum to string for JSON serialization
        supplier_dict["status"] = sample_supplier_data["status"].value
        supplier_dict["product_count"] = 5
        mock_service.get_supplier_with_product_count.return_value = supplier_dict

        supplier_id = sample_supplier_data["id"]
        response = client.get(f"/api/suppliers/{supplier_id}")

        assert response.status_code == 200
        data = response.json()
        assert data["product_count"] == 5
        assert data["name"] == "Test Supplier"

    @patch("app.api.supplier_routes.supplier_service")
    def test_get_supplier_not_found(self, mock_service, client):
        """Test supplier not found."""
        mock_service.get_supplier_with_product_count.return_value = None

        response = client.get(f"/api/suppliers/{uuid4()}")

        assert response.status_code == 404
        assert response.json()["detail"] == "Supplier not found"


class TestUpdateSupplier:
    """Tests for PUT /suppliers/{supplier_id} endpoint."""

    @patch("app.api.supplier_routes.supplier_service")
    def test_update_supplier_success(
        self, mock_service, client, sample_supplier_response
    ):
        """Test successful supplier update."""
        mock_service.update_supplier.return_value = sample_supplier_response

        supplier_id = sample_supplier_response.id
        response = client.put(
            f"/api/suppliers/{supplier_id}",
            json={"name": "Updated Name"},
        )

        assert response.status_code == 200
        mock_service.update_supplier.assert_called_once()

    @patch("app.api.supplier_routes.supplier_service")
    def test_update_supplier_not_found(self, mock_service, client):
        """Test update supplier not found."""
        mock_service.update_supplier.return_value = None

        response = client.put(
            f"/api/suppliers/{uuid4()}",
            json={"name": "Updated Name"},
        )

        assert response.status_code == 404

    @patch("app.api.supplier_routes.supplier_service")
    def test_update_supplier_validation_error(self, mock_service, client):
        """Test update with validation error from service layer."""
        mock_service.update_supplier.side_effect = ValueError("Invalid email format")

        response = client.put(
            f"/api/suppliers/{uuid4()}",
            json={"name": "Updated Name"},
        )

        assert response.status_code == 400
        assert "Invalid email format" in response.json()["detail"]


class TestDeleteSupplier:
    """Tests for DELETE /suppliers/{supplier_id} endpoint."""

    @patch("app.api.supplier_routes.supplier_service")
    def test_delete_supplier_success(self, mock_service, admin_client):
        """Test successful supplier deletion."""
        mock_service.delete_supplier.return_value = True

        response = admin_client.delete(f"/api/suppliers/{uuid4()}")

        assert response.status_code == 200
        assert response.json()["message"] == "Supplier deleted successfully"

    @patch("app.api.supplier_routes.supplier_service")
    def test_delete_supplier_not_found(self, mock_service, admin_client):
        """Test delete supplier not found."""
        mock_service.delete_supplier.return_value = False

        response = admin_client.delete(f"/api/suppliers/{uuid4()}")

        assert response.status_code == 404

    @patch("app.api.supplier_routes.supplier_service")
    def test_delete_supplier_with_active_products(self, mock_service, admin_client):
        """Test delete blocked when supplier has active products."""
        mock_service.delete_supplier.side_effect = ValueError(
            "Cannot delete supplier with active products"
        )

        response = admin_client.delete(f"/api/suppliers/{uuid4()}")

        assert response.status_code == 400
        assert "Cannot delete supplier with active products" in response.json()["detail"]

    def test_delete_supplier_forbidden_for_regular_user(self, client):
        """Test delete forbidden for regular users."""
        response = client.delete(f"/api/suppliers/{uuid4()}")

        assert response.status_code == 403


class TestGetSupplierProducts:
    """Tests for GET /suppliers/{supplier_id}/products endpoint."""

    @patch("app.api.supplier_routes.product_service")
    @patch("app.api.supplier_routes.supplier_service")
    def test_get_supplier_products_success(
        self,
        mock_supplier_service,
        mock_product_service,
        client,
        sample_supplier_response,
        sample_product_data,
    ):
        """Test successful retrieval of supplier products."""
        mock_supplier_service.get_supplier.return_value = sample_supplier_response

        product_response = ProductResponseDTO(**sample_product_data)
        mock_product_service.list_products.return_value = ProductListResponseDTO(
            items=[product_response],
            pagination=PaginationDTO(page=1, limit=20, total=1, pages=1),
            filters=ProductFilterDTO(supplier_id=sample_supplier_response.id),
        )

        response = client.get(
            f"/api/suppliers/{sample_supplier_response.id}/products"
        )

        assert response.status_code == 200
        data = response.json()
        assert len(data["items"]) == 1
        assert data["pagination"]["total"] == 1

    @patch("app.api.supplier_routes.supplier_service")
    def test_get_supplier_products_supplier_not_found(
        self, mock_supplier_service, client
    ):
        """Test products retrieval when supplier not found."""
        mock_supplier_service.get_supplier.return_value = None

        response = client.get(f"/api/suppliers/{uuid4()}/products")

        assert response.status_code == 404
        assert response.json()["detail"] == "Supplier not found"

    @patch("app.api.supplier_routes.product_service")
    @patch("app.api.supplier_routes.supplier_service")
    def test_get_supplier_products_with_pagination(
        self,
        mock_supplier_service,
        mock_product_service,
        client,
        sample_supplier_response,
    ):
        """Test products retrieval with pagination parameters."""
        mock_supplier_service.get_supplier.return_value = sample_supplier_response
        mock_product_service.list_products.return_value = ProductListResponseDTO(
            items=[],
            pagination=PaginationDTO(page=2, limit=10, total=0, pages=0),
            filters=ProductFilterDTO(supplier_id=sample_supplier_response.id),
        )

        response = client.get(
            f"/api/suppliers/{sample_supplier_response.id}/products?page=2&limit=10"
        )

        assert response.status_code == 200
        mock_product_service.list_products.assert_called_once()
        call_kwargs = mock_product_service.list_products.call_args.kwargs
        assert call_kwargs["page"] == 2
        assert call_kwargs["limit"] == 10


class TestExportVerificationExcel:
    """Tests for GET /suppliers/export/excel endpoint."""

    @patch("app.api.supplier_routes.supplier_service")
    def test_export_excel_success(self, mock_service, client):
        """Test successful Excel export returns correct content type."""
        # Create minimal valid Excel bytes using openpyxl
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Supplier Verification Data"
        ws.cell(row=1, column=1, value="Name")
        buffer = BytesIO()
        wb.save(buffer)
        mock_service.export_verification_excel.return_value = buffer.getvalue()

        response = client.get("/api/suppliers/export/excel")

        assert response.status_code == 200
        assert (
            response.headers["content-type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "attachment" in response.headers["content-disposition"]
        assert "supplier_verification_export_" in response.headers["content-disposition"]
        assert ".xlsx" in response.headers["content-disposition"]

    @patch("app.api.supplier_routes.supplier_service")
    def test_export_excel_with_filters(self, mock_service, client):
        """Test that filters are passed through to service."""
        from io import BytesIO
        from openpyxl import Workbook

        wb = Workbook()
        buffer = BytesIO()
        wb.save(buffer)
        mock_service.export_verification_excel.return_value = buffer.getvalue()

        response = client.get(
            "/api/suppliers/export/excel?certification_status=certified&search=test&sort_by=name&sort_order=desc"
        )

        assert response.status_code == 200
        mock_service.export_verification_excel.assert_called_once()
        call_kwargs = mock_service.export_verification_excel.call_args.kwargs
        assert call_kwargs["certification_status"] == "certified"
        assert call_kwargs["search"] == "test"
        assert call_kwargs["sort_by"] == "name"
        assert call_kwargs["sort_order"] == "desc"

    @patch("app.api.supplier_routes.supplier_service")
    def test_export_excel_parseable(self, mock_service, client):
        """Test that returned Excel file can be parsed by openpyxl."""
        from io import BytesIO
        from openpyxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Supplier Verification Data"
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Code")
        buffer = BytesIO()
        wb.save(buffer)
        mock_service.export_verification_excel.return_value = buffer.getvalue()

        response = client.get("/api/suppliers/export/excel")

        assert response.status_code == 200
        result_wb = load_workbook(BytesIO(response.content))
        assert "Supplier Verification Data" in result_wb.sheetnames
        ws = result_wb["Supplier Verification Data"]
        assert ws.cell(row=1, column=1).value == "Name"

    def test_export_excel_unauthenticated(self):
        """Test that unauthenticated requests are rejected."""
        test_app = FastAPI()
        test_app.include_router(router, prefix="/api/suppliers")
        unauth_client = TestClient(test_app)
        response = unauth_client.get("/api/suppliers/export/excel")

        assert response.status_code in [401, 403]

    @patch("app.api.supplier_routes.supplier_service")
    def test_export_excel_service_error(self, mock_service, client):
        """Test error handling when service raises exception."""
        mock_service.export_verification_excel.side_effect = Exception("DB error")

        response = client.get("/api/suppliers/export/excel")

        assert response.status_code == 500
        assert "Failed to export" in response.json()["detail"]
