"""Unit tests for the AI Data Extraction Service."""

import json
import tempfile
from decimal import Decimal
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from app.models.extraction_dto import (
    ExtractedProduct,
    ExtractionResult,
    HsCodeSuggestion,
    ImageOperation,
    ImageProcessingResult,
)
from app.services.extraction_service import ExtractionService


@pytest.fixture
def service():
    """Create a fresh ExtractionService instance for each test."""
    return ExtractionService()


@pytest.fixture
def mock_settings():
    """Create mock settings with API keys configured."""
    settings = MagicMock()
    settings.ANTHROPIC_API_KEY = "test-anthropic-key"
    settings.OPENAI_API_KEY = "test-openai-key"
    settings.REMOVEBG_API_KEY = "test-removebg-key"
    settings.EXTRACTION_AI_PROVIDER = "anthropic"
    settings.EXTRACTION_MAX_RETRIES = 3
    settings.EXTRACTION_TIMEOUT_SECONDS = 60
    return settings


@pytest.fixture
def mock_settings_no_ai():
    """Create mock settings without AI keys configured."""
    settings = MagicMock()
    settings.ANTHROPIC_API_KEY = None
    settings.OPENAI_API_KEY = None
    settings.REMOVEBG_API_KEY = None
    settings.EXTRACTION_AI_PROVIDER = "anthropic"
    settings.EXTRACTION_MAX_RETRIES = 3
    settings.EXTRACTION_TIMEOUT_SECONDS = 60
    return settings


class TestExtractionServiceInit:
    """Tests for ExtractionService initialization."""

    def test_service_initializes(self, service):
        """Test that service initializes correctly."""
        assert service is not None
        assert service._anthropic_client is None
        assert service._openai_client is None


class TestAIAvailability:
    """Tests for AI availability checking."""

    def test_ai_available_with_anthropic_key(self, service, mock_settings):
        """Test AI is available when Anthropic key is configured."""
        mock_settings.OPENAI_API_KEY = None
        with patch.object(service, "_settings", mock_settings):
            assert service._is_ai_available() is True

    def test_ai_available_with_openai_key(self, service, mock_settings):
        """Test AI is available when OpenAI key is configured."""
        mock_settings.ANTHROPIC_API_KEY = None
        with patch.object(service, "_settings", mock_settings):
            assert service._is_ai_available() is True

    def test_ai_not_available_without_keys(self, service, mock_settings_no_ai):
        """Test AI is not available when no keys are configured."""
        with patch.object(service, "_settings", mock_settings_no_ai):
            assert service._is_ai_available() is False


class TestPreferredProvider:
    """Tests for AI provider selection."""

    def test_prefers_anthropic_when_configured(self, service, mock_settings):
        """Test that Anthropic is preferred when configured."""
        with patch.object(service, "_settings", mock_settings):
            assert service._get_preferred_ai_provider() == "anthropic"

    def test_prefers_openai_when_anthropic_unavailable(self, service, mock_settings):
        """Test fallback to OpenAI when Anthropic is unavailable."""
        mock_settings.ANTHROPIC_API_KEY = None
        mock_settings.EXTRACTION_AI_PROVIDER = "anthropic"
        with patch.object(service, "_settings", mock_settings):
            assert service._get_preferred_ai_provider() == "openai"

    def test_returns_none_when_no_provider(self, service, mock_settings_no_ai):
        """Test returns 'none' when no provider is available."""
        with patch.object(service, "_settings", mock_settings_no_ai):
            assert service._get_preferred_ai_provider() == "none"


class TestParseExtractionResponse:
    """Tests for parsing AI extraction responses."""

    def test_parses_valid_json_response(self, service):
        """Test parsing a valid JSON response."""
        response = json.dumps({
            "sku": "TEST-001",
            "name": "Test Product",
            "description": "A test product",
            "price_fob_usd": 10.50,
            "moq": 100,
        })
        result = service._parse_extraction_response(response)

        assert result.sku == "TEST-001"
        assert result.name == "Test Product"
        assert result.description == "A test product"
        assert result.price_fob_usd == Decimal("10.50")
        assert result.moq == 100
        assert result.confidence_score == 1.0  # All 5 fields present

    def test_parses_partial_json_response(self, service):
        """Test parsing a response with missing fields."""
        response = json.dumps({"sku": "TEST-002", "name": "Partial Product"})
        result = service._parse_extraction_response(response)

        assert result.sku == "TEST-002"
        assert result.name == "Partial Product"
        assert result.price_fob_usd is None
        assert result.confidence_score == 0.4  # 2 of 5 fields

    def test_handles_invalid_json(self, service):
        """Test handling of invalid JSON response."""
        response = "This is not valid JSON"
        result = service._parse_extraction_response(response)

        assert result.sku is None
        assert result.name is None
        assert result.confidence_score == 0.0

    def test_extracts_json_from_text(self, service):
        """Test extracting JSON embedded in text."""
        response = 'Here is the result: {"sku": "EMB-001", "name": "Embedded"} end.'
        result = service._parse_extraction_response(response)

        assert result.sku == "EMB-001"
        assert result.name == "Embedded"

    def test_preserves_source_page(self, service):
        """Test that source page is preserved in result."""
        response = json.dumps({"sku": "PAGE-001"})
        result = service._parse_extraction_response(response, source_page=5)

        assert result.source_page == 5


class TestExtractProductData:
    """Tests for product data extraction."""

    def test_returns_empty_when_ai_unavailable(self, service, mock_settings_no_ai):
        """Test graceful fallback when AI is unavailable."""
        with patch.object(service, "_settings", mock_settings_no_ai):
            result = service.extract_product_data("Some product text")

            assert result.confidence_score == 0.0
            assert result.raw_text == "Some product text"

    @patch("app.services.extraction_service.ExtractionService._extract_with_anthropic")
    def test_uses_anthropic_when_preferred(
        self, mock_extract, service, mock_settings
    ):
        """Test that Anthropic is used when it's the preferred provider."""
        mock_extract.return_value = json.dumps({"sku": "ANT-001", "name": "Anthropic"})

        with patch.object(service, "_settings", mock_settings):
            result = service.extract_product_data("Test content")

            mock_extract.assert_called_once()
            assert result.sku == "ANT-001"


class TestProcessExcel:
    """Tests for Excel file processing."""

    def test_handles_missing_file(self, service):
        """Test handling of missing file."""
        products, errors = service.process_excel("/nonexistent/file.xlsx")

        assert len(products) == 0
        assert len(errors) == 1
        assert "File not found" in errors[0]

    def test_processes_excel_with_standard_columns(self, service):
        """Test processing Excel with standard column names."""
        from openpyxl import Workbook

        # Create a test Excel file
        wb = Workbook()
        ws = wb.active
        ws.append(["SKU", "Name", "Price", "MOQ", "Description"])
        ws.append(["TEST-001", "Product One", "10.00", "100", "First product"])
        ws.append(["TEST-002", "Product Two", "20.00", "50", "Second product"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            products, errors = service.process_excel(temp_path)

            assert len(products) == 2
            assert len(errors) == 0
            assert products[0].sku == "TEST-001"
            assert products[0].name == "Product One"
            assert products[0].price_fob_usd == Decimal("10.00")
            assert products[0].moq == 100
            assert products[1].sku == "TEST-002"
        finally:
            Path(temp_path).unlink()

    def test_processes_excel_with_alternative_columns(self, service):
        """Test processing Excel with alternative column names."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["Reference", "Product", "FOB", "Minimum"])
        ws.append(["ALT-001", "Alt Product", "$15.50", "200"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            products, errors = service.process_excel(temp_path)

            assert len(products) == 1
            assert products[0].sku == "ALT-001"
            assert products[0].name == "Alt Product"
            assert products[0].price_fob_usd == Decimal("15.50")
            assert products[0].moq == 200
        finally:
            Path(temp_path).unlink()

    def test_skips_empty_rows(self, service):
        """Test that empty rows are skipped."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["SKU", "Name"])
        ws.append(["SKIP-001", "Valid Product"])
        ws.append([None, None])  # Empty row
        ws.append(["SKIP-002", "Another Product"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            products, errors = service.process_excel(temp_path)

            assert len(products) == 2
        finally:
            Path(temp_path).unlink()


class TestProcessPdf:
    """Tests for PDF file processing."""

    def test_handles_missing_file(self, service):
        """Test handling of missing file."""
        products, errors = service.process_pdf("/nonexistent/file.pdf")

        assert len(products) == 0
        assert len(errors) == 1
        assert "File not found" in errors[0]


class TestProcessImage:
    """Tests for image file processing."""

    def test_handles_missing_file(self, service):
        """Test handling of missing file."""
        product, errors = service.process_image("/nonexistent/image.jpg")

        assert product is None
        assert len(errors) == 1
        assert "File not found" in errors[0]


class TestProcessBatch:
    """Tests for batch file processing."""

    def test_processes_empty_batch(self, service):
        """Test processing an empty batch."""
        result = service.process_batch([])

        assert isinstance(result, ExtractionResult)
        assert result.total_extracted == 0
        assert len(result.errors) == 0

    def test_handles_unsupported_file_types(self, service):
        """Test handling of unsupported file types."""
        result = service.process_batch(["/some/file.txt", "/another/file.doc"])

        assert result.total_extracted == 0
        assert len(result.warnings) >= 2

    def test_continues_on_error(self, service):
        """Test that batch processing continues after errors."""
        from openpyxl import Workbook

        # Create one valid file
        wb = Workbook()
        ws = wb.active
        ws.append(["SKU", "Name"])
        ws.append(["BATCH-001", "Valid Product"])

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb.save(f.name)
            temp_path = f.name

        try:
            result = service.process_batch(
                ["/nonexistent.pdf", temp_path, "/another_nonexistent.jpg"]
            )

            assert result.total_extracted == 1
            assert len(result.errors) >= 2  # Two missing files
        finally:
            Path(temp_path).unlink()

    def test_includes_ai_warning_when_unavailable(self, service, mock_settings_no_ai):
        """Test that a warning is added when AI is unavailable."""
        with patch.object(service, "_settings", mock_settings_no_ai):
            result = service.process_batch([])

            assert any("AI services unavailable" in w for w in result.warnings)


class TestSuggestHsCode:
    """Tests for HS code suggestion."""

    def test_returns_fallback_when_ai_unavailable(self, service, mock_settings_no_ai):
        """Test graceful fallback when AI is unavailable."""
        with patch.object(service, "_settings", mock_settings_no_ai):
            result = service.suggest_hs_code("Test product description")

            assert isinstance(result, HsCodeSuggestion)
            assert result.code == "9999.99.99"
            assert result.confidence_score == 0.0

    @patch("app.services.extraction_service.ExtractionService._get_anthropic_client")
    def test_parses_valid_hs_code_response(
        self, mock_client, service, mock_settings
    ):
        """Test parsing a valid HS code response."""
        mock_anthropic = MagicMock()
        mock_message = MagicMock()
        mock_message.content = [
            MagicMock(
                text=json.dumps({
                    "code": "6204.42.00",
                    "description": "Women's dresses of cotton",
                    "confidence_score": 0.85,
                    "reasoning": "Based on cotton dress description",
                })
            )
        ]
        mock_anthropic.messages.create.return_value = mock_message
        mock_client.return_value = mock_anthropic

        with patch.object(service, "_settings", mock_settings):
            result = service.suggest_hs_code("Cotton dress for women")

            assert result.code == "6204.42.00"
            assert result.confidence_score == 0.85


class TestRemoveBackground:
    """Tests for background removal."""

    def test_returns_original_when_api_key_missing(self, service, mock_settings_no_ai):
        """Test that original URL is returned when API key is missing."""
        with patch.object(service, "_settings", mock_settings_no_ai):
            result = service.remove_background("https://example.com/image.jpg")

            assert isinstance(result, ImageProcessingResult)
            assert result.original_url == "https://example.com/image.jpg"
            assert result.processed_url == "https://example.com/image.jpg"
            assert result.operation == ImageOperation.REMOVE_BG

    @patch("httpx.Client")
    def test_calls_removebg_api(self, mock_httpx, service, mock_settings):
        """Test that RemoveBG API is called correctly."""
        mock_response = MagicMock()
        mock_response.status_code = 200
        mock_response.content = b"fake_image_data"

        mock_client_instance = MagicMock()
        mock_client_instance.post.return_value = mock_response
        mock_client_instance.__enter__ = MagicMock(return_value=mock_client_instance)
        mock_client_instance.__exit__ = MagicMock(return_value=False)
        mock_httpx.return_value = mock_client_instance

        with patch.object(service, "_settings", mock_settings):
            result = service.remove_background("https://example.com/image.jpg")

            assert result.operation == ImageOperation.REMOVE_BG
            assert result.processed_url.startswith("data:image/png;base64,")


class TestResizeImage:
    """Tests for image resizing."""

    def test_handles_missing_file(self, service):
        """Test handling of missing file."""
        result, error = service.resize_image("/nonexistent/image.jpg", 100, 100)

        assert result is None
        assert error is not None

    def test_resizes_valid_image(self, service):
        """Test resizing a valid image."""
        from PIL import Image as PILImage

        # Create a test image
        img = PILImage.new("RGB", (200, 200), color="red")

        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as f:
            img.save(f.name)
            temp_path = f.name

        try:
            result, error = service.resize_image(temp_path, 100, 100)

            assert error is None
            assert result is not None
            assert result.startswith("data:image/")
        finally:
            Path(temp_path).unlink()


class TestFindHigherQualityImage:
    """Tests for higher quality image search."""

    def test_returns_none_placeholder(self, service):
        """Test that placeholder returns None."""
        result = service.find_higher_quality_image("https://example.com/image.jpg")

        assert result is None


class TestExtractionDTOs:
    """Tests for extraction DTOs."""

    def test_extracted_product_defaults(self):
        """Test ExtractedProduct default values."""
        product = ExtractedProduct()

        assert product.sku is None
        assert product.name is None
        assert product.image_urls == []
        assert product.confidence_score == 0.0

    def test_extraction_result_defaults(self):
        """Test ExtractionResult default values."""
        result = ExtractionResult()

        assert result.products == []
        assert result.total_extracted == 0
        assert result.errors == []
        assert result.warnings == []

    def test_hs_code_suggestion_validation(self):
        """Test HsCodeSuggestion field validation."""
        suggestion = HsCodeSuggestion(
            code="1234.56.78",
            description="Test description",
            confidence_score=0.9,
        )

        assert suggestion.code == "1234.56.78"
        assert suggestion.confidence_score == 0.9

    def test_image_processing_result(self):
        """Test ImageProcessingResult creation."""
        result = ImageProcessingResult(
            original_url="https://example.com/original.jpg",
            processed_url="https://example.com/processed.jpg",
            operation=ImageOperation.RESIZE,
        )

        assert result.operation == ImageOperation.RESIZE
