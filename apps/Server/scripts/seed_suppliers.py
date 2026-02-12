#!/usr/bin/env python3
"""Seed script for suppliers from Canton Fair master directory Excel."""

import re
import sys
import os
from typing import Optional

# Add the parent directory to the path so we can import app modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.repository.kompass_repository import supplier_repository

# Canonical names for the 25 known suppliers
KNOWN_SUPPLIERS = [
    "BWBYONE",
    "BATH STORE-TAUSU",
    "CONRAZZO",
    "CU MATERIALS",
    "DHF",
    "FOSHAN SHISUO",
    "GEORGE",
    "HONGYU",
    "HUAYI",
    "JINGDA",
    "JVK",
    "LAYASDUN",
    "LEIZI",
    "LUXDREAM",
    "MAYORISTA COLOMBIA",
    "MEXY TECH",
    "NTFT YIFUYUAN",
    "PA HOME",
    "PINLSLON",
    "SENCHUAN",
    "SOTENG",
    "VAN-RUBEN",
    "WEIRE",
    "WIREKING",
    "WUXI KAIZE",
]

# Default Excel path relative to the repo root
DEFAULT_EXCEL_PATH = os.path.join(
    "Requirements_Gathering",
    "Sourcing",
    "Data",
    "PROVEEDORES - CATALOGOS",
    "CANTON FAIR _OCT _2025",
    "DIRECTORIO PROVEEDORES CANTON FAIR _OCT_2025.xlsx",
)


def normalize_name(name: str) -> str:
    """Normalize a supplier name for comparison: strip, uppercase, collapse spaces."""
    return re.sub(r"\s+", " ", name.strip().upper())


def find_matching_supplier(excel_name: str, known_list: list) -> Optional[str]:
    """Find a matching known supplier for an Excel name.

    Tries exact match first, then startswith/containment.
    Returns the canonical known supplier name or None.
    """
    normalized_excel = normalize_name(excel_name)
    normalized_known = {normalize_name(k): k for k in known_list}

    # Exact match
    if normalized_excel in normalized_known:
        return normalized_known[normalized_excel]

    # Startswith or containment match
    for norm_name, canonical in normalized_known.items():
        if normalized_excel.startswith(norm_name) or norm_name.startswith(
            normalized_excel
        ):
            return canonical
        if norm_name in normalized_excel or normalized_excel in norm_name:
            return canonical

    return None


def seed_suppliers(excel_path: str) -> dict:
    """Seed suppliers from the Canton Fair master directory Excel.

    Args:
        excel_path: Path to the Canton Fair directory Excel file.

    Returns:
        Dict with 'mappings' ({supplier_name: uuid_string}),
        'created' count, 'skipped' count, and 'not_found_in_excel' list.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR [SeedSuppliers]: openpyxl not installed. Run: pip install openpyxl")
        return {"mappings": {}, "created": 0, "skipped": 0, "not_found_in_excel": KNOWN_SUPPLIERS[:]}

    if not os.path.exists(excel_path):
        print(f"ERROR [SeedSuppliers]: Excel file not found: {excel_path}")
        return {"mappings": {}, "created": 0, "skipped": 0, "not_found_in_excel": KNOWN_SUPPLIERS[:]}

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Try to find the sheet
    sheet_name = "BASE MASTER"
    if sheet_name not in wb.sheetnames:
        # Fall back to first sheet
        print(
            f"  WARN [SeedSuppliers]: Sheet '{sheet_name}' not found, "
            f"using first sheet: '{wb.sheetnames[0]}'"
        )
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Read header row to find column indices
    headers = {}
    header_map = {
        "proveedor": "proveedor",
        "contacto": "contacto",
        "email": "email",
        "ciudad": "ciudad",
        "pagina web": "pagina_web",
        "productos": "productos",
    }
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
        for cell in row:
            if cell.value:
                cell_lower = str(cell.value).strip().lower()
                for key, mapped in header_map.items():
                    if key in cell_lower:
                        headers[mapped] = cell.column - 1  # 0-indexed
                        break

    if "proveedor" not in headers:
        print("ERROR [SeedSuppliers]: 'Proveedor' column not found in Excel headers")
        wb.close()
        return {"mappings": {}, "created": 0, "skipped": 0, "not_found_in_excel": KNOWN_SUPPLIERS[:]}

    mappings = {}
    created = 0
    skipped = 0
    seen_suppliers = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= headers["proveedor"]:
            continue

        proveedor_val = row[headers["proveedor"]]
        if not proveedor_val or not str(proveedor_val).strip():
            continue

        excel_name = str(proveedor_val).strip()
        matched_name = find_matching_supplier(excel_name, KNOWN_SUPPLIERS)

        if not matched_name or matched_name in seen_suppliers:
            continue

        seen_suppliers.add(matched_name)

        # Check if supplier already exists in DB
        existing = supplier_repository.get_by_name(matched_name)
        if existing:
            mappings[matched_name] = str(existing["id"])
            print(
                f"  SKIP [Supplier]: '{matched_name}' already exists ({existing['id']})"
            )
            skipped += 1
            continue

        # Extract optional fields from row
        contact_name = None
        contact_email = None
        city = None
        website = None
        notes = None

        if "contacto" in headers and len(row) > headers["contacto"]:
            val = row[headers["contacto"]]
            if val:
                contact_name = str(val).strip() or None

        if "email" in headers and len(row) > headers["email"]:
            val = row[headers["email"]]
            if val:
                contact_email = str(val).strip() or None

        if "ciudad" in headers and len(row) > headers["ciudad"]:
            val = row[headers["ciudad"]]
            if val:
                city = str(val).strip() or None

        if "pagina_web" in headers and len(row) > headers["pagina_web"]:
            val = row[headers["pagina_web"]]
            if val:
                website = str(val).strip() or None

        if "productos" in headers and len(row) > headers["productos"]:
            val = row[headers["productos"]]
            if val:
                notes = str(val).strip() or None

        result = supplier_repository.create(
            name=matched_name,
            contact_name=contact_name,
            contact_email=contact_email,
            city=city,
            country="China",
            website=website,
            notes=notes,
            status="active",
        )

        if result:
            mappings[matched_name] = str(result["id"])
            print(f"  CREATE [Supplier]: '{matched_name}' ({result['id']})")
            created += 1
        else:
            print(f"  ERROR [Supplier]: Failed to create '{matched_name}'")

    wb.close()

    not_found = [s for s in KNOWN_SUPPLIERS if s not in seen_suppliers]

    return {
        "mappings": mappings,
        "created": created,
        "skipped": skipped,
        "not_found_in_excel": not_found,
    }


if __name__ == "__main__":
    # Accept custom Excel path from CLI or use default
    if len(sys.argv) > 1:
        path = sys.argv[1]
    else:
        # Resolve relative to repo root (3 levels up from scripts/)
        repo_root = os.path.dirname(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        )
        path = os.path.join(repo_root, DEFAULT_EXCEL_PATH)

    print(f"INFO [SeedSuppliers]: Starting supplier seeding from: {path}")
    result = seed_suppliers(path)
    print(
        f"\nINFO [SeedSuppliers]: Done. "
        f"Created: {result['created']}, Skipped: {result['skipped']}"
    )
    if result["not_found_in_excel"]:
        print(
            f"  Not found in Excel: {', '.join(result['not_found_in_excel'])}"
        )
